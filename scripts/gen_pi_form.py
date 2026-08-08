"""
TMT MASTER PI INTAKE -> ClaimReach form data

Reads TMT_Master_PI_Intake.xlsx and emits the seed migration. The spreadsheet is
the authored source of truth for CONTENT; this file only maps it onto keys and
conditions. Nothing here writes a question. If wording is wrong, it is wrong in
the spreadsheet and gets fixed there, then regenerated.

  python3 scripts/gen_pi_form.py TMT_Master_PI_Intake.xlsx > supabase/migrations/0077_tmt_pi_form.sql

Two things this has to get right or the platform breaks quietly.

1. ROUTING KEYS. The disposition engine matches on specific field ids
   (injured, treatment, bills, attorney, settled, date, willing, agent_read).
   Core questions are mapped ONTO those ids rather than given new ones, so the
   existing SIGN/REFER/DQ logic keeps working against the new script.

2. THE ROUTER. Core Q5 sets case_subtype, and every branch question carries
   showIf case_subtype = its branch. The agent is never asked to decide which
   questions to ask: they mark what the caller described and the branch follows.
"""
import json, sys, re
import openpyxl

# Core Q5 option label -> internal value -> the EXACT string Lexamica expects.
# Note the label drift: the script says "Dog Bite Injury", Lexamica wants
# "Dog Bite Injuries". A mismatch there lands the case in Other and it is never
# referred out, so the two are stated together here and nowhere else.
SUBTYPES = [
    ("Motor Vehicle Accident",                "mva",        "Motor Vehicle Accident",       "03 Motor Vehicle"),
    ("Personal Injury (general or premises)", "general",    "Personal Injury",              "04 Personal Injury General"),
    ("Dog Bite Injury",                       "dogbite",    "Dog Bite Injuries",            "05 Dog Bite"),
    ("Workplace Injury",                      "workplace",  "Workplace Injuries",           "06 Workplace Injury"),
    ("Workers' Compensation",                 "workcomp",   "Workers' Compensation",        "07 Workers Comp"),
    ("Pedestrian Injury",                     "pedestrian", "Pedestrian Injuries",          "08 Pedestrian"),
    ("Commercial Property Injury",            "commprop",   "Commercial Property Injuries", "09 Commercial Property"),
    ("Construction Accident",                 "construct",  "Construction Accidents",       "10 Construction"),
    ("Medical Malpractice",                   "medmal",     "Medical Malpractice",          "11 Medical Malpractice"),
    ("Product Liability",                     "prodliab",   "Product Liability",            "12 Product Liability"),
    ("Nursing Home Injury",                   "nursing",    "Nursing Home Injuries",        "13 Nursing Home"),
    ("Not a case type we handle",             "referout",   "Personal Injury",              "14 Refer Out"),
]

# Core question number -> field id. Ids that already exist in the engine are
# reused deliberately; see note 1 above.
CORE_KEYS = {
    "1": "authority", "2": "ip_adult", "3": "poa", "4": "attorney",
    "5": "case_subtype", "6": "what_happened", "7": "date", "8": "discovery_date",
    "9": "incident_city", "10": "incident_state", "11": "incident_place",
    "12": "injured", "13": "injuries", "14": "symptoms_ongoing",
    "15": "treatment", "16": "level_of_care", "17": "willing", "18": "willing_more",
    "19": "bills", "20": "responsible_party", "21": "incident_report",
    "22": "responders", "23": "witnesses", "24": "photos",
    "25": "settled", "26": "health_insurance", "27": "sol_status",
    "28": "agent_read", "29": "case_manager_notes",
}

# Branch questions that ARE engine keys. Given the engine id, not a prefixed one,
# so MVA fault and insurance keep driving disposition exactly as before.
BRANCH_KEYS = {
    "M2": "fault", "M3": "police_report", "M4": "citations", "M5": "commercial",
    "M6": "collision_type", "M7": "ins_other", "M8": "ins_own", "M9": "ins_uim",
    "M1": "role", "M12": "others_in_vehicle",
    "G2": "presence",
}

# Answer values for the routing keys. These are matched literally by the engine,
# so they are stated here rather than slugified from the label.
FIXED_VALUES = {
    "authority":  ["self", "alive", "deceased"],
    "attorney":   ["no", "yes"],
    "injured":    ["yes", "no"],
    "treatment":  ["still", "finished", "stopped", "never"],
    "bills":      ["none", "under_10k", "10k_50k", "over_50k", "unknown"],
    "willing":    ["yes", "no"],
    "willing_more": ["yes", "no", "unsure"],
    "settled":    ["no", "yes"],
    "symptoms_ongoing": ["yes", "no"],
    "poa":        ["yes", "no"],
    "ip_adult":   ["yes", "no"],
    "presence":   ["yes", "no", "unsure"],
    "fault":      ["other", "shared", "caused"],
    "police_report": ["yes", "no", "unsure"],
    "ins_other":  ["yes", "no", "unsure"],
    "ins_own":    ["yes", "no", "unsure"],
    "ins_uim":    ["yes", "no", "unsure"],
    "role":       ["driver", "passenger"],
    "injuries":   ["neck_back","strain","lig_strain","anxiety","head","broken",
                   "lig_tear","internal","laceration","burns","scarring",
                   "spinal","amputation","bedsores","death"],
    "agent_read": ["sign_high_value", "sign", "refer", "escalate"],
    "sol_status": ["clear", "under_6mo", "expired", "government"],
    # M5. The engine fires the CMV modifier on commercial == "yes", so the
    # commercial option must BE "yes". Slugifying the label to
    # "commercial_vehicle" would have left the modifier silently never firing.
    "commercial": ["no", "yes", "government", "rideshare", "unknown"],
    "citations":  ["other", "caller", "none", "unsure"],
    "collision_type": ["rear_end", "head_on", "side", "rollover", "multi", "hit_run", "other"],
    "others_in_vehicle": ["no", "others_none_injured", "others_injured", "unsure"],
    "level_of_care": ["ambulance", "er_same_day", "hosp_short", "hosp_long",
                      "surgery_done", "surgery_recommended", "pt_injections",
                      "wound_care", "none"],
}

KIND = {
    "Radio": "select", "Dropdown": "select", "Multi-Select": "multiselect",
    "Single Line": "text", "Paragraph": "longtext", "Date": "date",
}

def slug(s):
    s = re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")
    return s[:40] or "opt"

def opts_for(fid, raw):
    labels = [o.strip() for o in (raw or "").split("\n") if o.strip()]
    labels = [l for l in labels if not l.startswith("[")]
    if not labels:
        return None
    if fid in FIXED_VALUES and len(FIXED_VALUES[fid]) == len(labels):
        vals = FIXED_VALUES[fid]
    else:
        vals, seen = [], set()
        for l in labels:
            v = slug(l); n = v; i = 2
            while n in seen: n = f"{v}_{i}"; i += 1
            seen.add(n); vals.append(n)
    return [{"value": v, "label": l} for v, l in zip(vals, labels)]

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    fields = []

    def add(fid, kind, script, raw_opts, note, show=None, origin="spine"):
        f = {"id": fid, "scope": "lead", "kind": kind,
             "label": script if script and not script.startswith("[") else fid,
             "origin": origin, "locked": True}
        if script and not script.startswith("["):
            f["script"] = script
        if note: f["agentNote"] = note
        if kind in ("longtext",): f["multiline"] = True
        ch = opts_for(fid, raw_opts)
        if ch:
            f["choices"] = ch
            f["options"] = [c["label"] for c in ch]
        if show: f["showIf"] = show
        fields.append(f)

    # ---- Core Screen: asked on every call
    ws = wb["02 Core Screen"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        c = [("" if x is None else str(x).strip()) for x in r]
        num = c[0]
        if num not in CORE_KEYS: continue
        fid = CORE_KEYS[num]
        kind = KIND.get(c[1], "text")
        if fid == "case_subtype":
            ch = [{"value": v, "label": lbl} for lbl, v, _, _ in SUBTYPES]
            fields.append({
                "id": "case_subtype", "scope": "lead", "kind": "select",
                "label": "Case type", "origin": "spine", "locked": True,
                "routingKey": True,
                "agentNote": c[4] or "This routes the call. Mark what the caller described; the matching questions follow automatically.",
                "choices": ch, "options": [x["label"] for x in ch],
            })
            continue
        show = None
        if fid == "poa":       show = {"match":"all","rules":[{"fieldId":"authority","op":"any_of","values":["alive","deceased"]}]}
        if fid == "injuries":  show = {"match":"all","rules":[{"fieldId":"injured","op":"is","value":"yes"}]}
        if fid == "symptoms_ongoing": show = {"match":"all","rules":[{"fieldId":"injured","op":"is","value":"yes"}]}
        if fid in ("treatment","level_of_care","bills"): show = {"match":"all","rules":[{"fieldId":"injured","op":"is","value":"yes"}]}
        if fid == "willing":   show = {"match":"all","rules":[{"fieldId":"treatment","op":"is","value":"never"}]}
        if fid == "willing_more": show = {"match":"all","rules":[{"fieldId":"treatment","op":"any_of","values":["finished","stopped"]}]}
        if fid == "discovery_date": show = {"match":"all","rules":[{"fieldId":"case_subtype","op":"any_of","values":["medmal","prodliab","nursing"]}]}
        add(fid, kind, c[2], c[3], c[4], show)

    # ---- Branches: only the one Q5 selected
    for label, val, _lex, tab in SUBTYPES:
        if tab not in wb.sheetnames: continue
        base = {"match":"all","rules":[{"fieldId":"case_subtype","op":"is","value":val}]}
        block = None
        for r in wb[tab].iter_rows(min_row=2, values_only=True):
            c = [("" if x is None else str(x).strip()) for x in r]
            num = c[0]
            if not num or num == "#": continue
            # Section banners carry the block gating for tab 04
            if not c[2] or len(num) > 6:
                m = re.search(r"ask if G1 is (.+?)\)", num)
                if m:
                    t = m.group(1).lower()
                    if "fall" in t or "hazard" in t: block = "fall_or_hazard"
                    elif "attack" in t: block = "assault"
                    elif "intoxicated" in t: block = "dram_shop"
                continue
            if c[3] == "[Not a field]" or c[3] == "[Do not ask this question]": continue
            fid = BRANCH_KEYS.get(num) or f"{val}_{slug(c[2])[:32]}"
            show = base
            if tab == "04 Personal Injury General" and block and num not in ("G1","G2","G3","G4","G5","G6","G7","G8"):
                show = {"match":"all","rules":[
                    {"fieldId":"case_subtype","op":"is","value":val},
                    {"fieldId":"general_incident_kind","op":"any_of","values":[block]},
                ]}
            if num == "G1": fid = "general_incident_kind"
            add(fid, KIND.get(c[1], "text"), c[2], c[3], c[4], show)

    return fields

if __name__ == "__main__":
    fields = main(sys.argv[1])
    ask = [f["id"] for f in fields]
    lex = {v: lx for _l, v, lx, _t in SUBTYPES}
    def sql_lit(obj):
        return "'" + json.dumps(obj).replace("'", "''") + "'"

    print("-- Generated by scripts/gen_pi_form.py. Do not hand edit; regenerate.")
    print("-- Source: TMT_Master_PI_Intake.xlsx")
    print("-- %d fields. An agent sees roughly 38 on any one call: the core screen" % len(fields))
    print("-- plus the single branch that case_subtype selected.")
    print("delete from intake_forms where claim_type = 'prem' and firm_id is null and campaign_id is null;")
    print("insert into intake_forms (firm_id, campaign_id, claim_type, name, description, status, version, fields, ask_order) values")
    print("  (null, null, 'prem', 'TMT Master PI Intake',")
    print("   'Core screen plus one branch, selected by case_subtype. Generated from TMT_Master_PI_Intake.xlsx.',")
    print("   'published', 20,")
    print("   " + sql_lit(fields) + "::jsonb,")
    print("   " + sql_lit(ask) + "::jsonb);")
    print()
    print("-- Subtype -> Lexamica PracticeArea, for reference. The app reads this from code.")
    for k, v in lex.items():
        print(f"--   {k:12} -> {v}")
